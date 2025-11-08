# importy
import psycopg2
import mariadb

from psycopg2.extras import RealDictCursor
from openpyxl import Workbook

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from pprint import pprint

db_type = 'postgres'

# sql config postgres
config = {
    'postgres': {
        'host': 'localhost',
        'user': 'postgres',
        'password': '123SQL123+',
        'dbname': 'classicmodels',
    },

    'mariadb': {
        'user': 'guest',
        'password': 'ctu-relational',
        'host': 'relational.fel.cvut.cz',
        'port': 3306,
        'database': 'classicmodels',
    }
}

queries = {
    'all': "SELECT * FROM products;",

    'scale': """
             SELECT productscale,
                    COUNT(*) AS product_count
             FROM products
             GROUP BY productscale
             ORDER BY product_count DESC;""",

    'line': """
            SELECT products.productline,
                   SUM(orderdetails.quantityordered) AS total_quantity
            FROM products
                     JOIN orderdetails ON products.productcode = orderdetails.productcode
            GROUP BY products.productline
            ORDER BY total_quantity DESC;"""
}


def get_connection(connection_type: str, inner_config: dict):
    if connection_type == 'postgres':
        return psycopg2.connect(**inner_config)
    elif connection_type == 'mariadb':
        return mariadb.connect(**inner_config)


def get_data(connection, query) -> dict:
    if type(connection) == psycopg2.extensions.connection:
        cursor = connection.cursor(cursor_factory=RealDictCursor)
    elif type(connection) == mariadb.connections.Connection:
        cursor = connection.cursor(dictionary=True)

    cursor.execute(query)

    return cursor.fetchall()


def get_whole_data(inner_queries):
    connection = None
    data = {}
    try:
        connection = get_connection(db_type, config[db_type])
        for q in inner_queries:
            current_query = inner_queries[q]
            data[q] = get_data(connection, current_query)
    finally:
        if connection is not None:
            connection.close()

    return data


def generate_excel(inner_data):
    wb = Workbook()

    for no, key in enumerate(inner_data):
        if no == 0:
            ws = wb.active
            ws.title = key
        else:
            ws = wb.create_sheet(title=key)

        data_list = inner_data[key]
        ws.append(list(data_list[0].keys()))
        for row in data_list:
            ws.append(list(row.values()))

        data_range = ws.dimensions
        tab = Table(displayName=key, ref=data_range)

        style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        tab.tableStyleInfo = style

        ws.add_table(tab)

        for i in range(1, ws.max_column + 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 25

    wb.save('Report.xlsx')


def main():
    data = get_whole_data(queries)
    generate_excel(data)


main()